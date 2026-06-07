# ============================================================
#  Bottom-Up Beta Calculator  |  v1.0
#  Streamlit app — fetches Yahoo Finance historical closing
#  prices and exports a formatted Excel workbook.
# ============================================================

import io
import warnings
from datetime import date, datetime

import pandas as pd
import streamlit as st
import yfinance as yf
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bottom-Up Beta Calculator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# Global CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
<style>
[data-testid="stAppViewContainer"] { background: #010B13; }
[data-testid="stHeader"] { background: transparent; }

.card {
    background: #111827;
    border-radius: 12px;
    padding: 22px 26px 18px;
    margin-bottom: 20px;
    border: 1px solid #374151;
    border-top: 3px solid #f59e0b;
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.4);
}

.card-title {
    font-size: 15px;
    font-weight: 700;
    color: #f9fafb;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 16px;
    padding-bottom: 10px;
    border-bottom: 1px solid #374151;
}

.note {
    background: linear-gradient(135deg, #1e293b, #0f172a);
    border: 1px solid #334155;
    border-left: 5px solid #3b82f6;
    border-radius: 0 8px 8px 0;
    padding: 14px 18px;
    font-size: 13px;
    color: #f8fafc;
    line-height: 1.65;
    margin-bottom: 20px;
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.35);
}
code {
    background: #e8f0fb;
    padding: 1px 5px;
    border-radius: 3px;
    font-size: 12px;
}

/* All widget labels */
[data-testid="stWidgetLabel"] {
    color: #ffffff !important;
    font-weight: 600 !important;
}
</style>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
INDICES: dict[str, str] = {
    "S&P 500 (USA)":              "^GSPC",
    "NASDAQ Composite (USA)":     "^IXIC",
    "Dow Jones Industrial (USA)": "^DJI",
    "NYSE Composite (USA)":       "^NYA",
    "Russell 2000 (USA)":         "^RUT",
    "FTSE 100 (UK)":              "^FTSE",
    "DAX (Germany)":              "^GDAXI",
    "CAC 40 (France)":            "^FCHI",
    "Euro Stoxx 50 (Europe)":     "^STOXX50E",
    "Nikkei 225 (Japan)":         "^N225",
    "Hang Seng (Hong Kong)":      "^HSI",
    "Shanghai Composite (China)": "000001.SS",
    "BSE SENSEX (India)":         "^BSESN",
    "NIFTY 50 (India)":           "^NSEI",
    "ASX 200 (Australia)":        "^AXJO",
    "KOSPI (South Korea)":        "^KS11",
    "TSX Composite (Canada)":     "^GSPTSE",
    "Bovespa (Brazil)":           "^BVSP",
    "MSCI World ETF (URTH)":      "URTH",
    "── Custom Ticker ──":        "__CUSTOM__",
}

INTERVAL: dict[str, str] = {"Daily": "1d", "Weekly": "1wk", "Monthly": "1mo"}
EST_PTS: dict[str, int]   = {"Daily": 252, "Weekly": 52,    "Monthly": 12}

# ─────────────────────────────────────────────────────────────────────────────
# Session-state helpers
# ─────────────────────────────────────────────────────────────────────────────
if "companies" not in st.session_state:
    st.session_state.companies: list[str] = ["", "", ""]


def _sync_inputs() -> None:
    """Pull current text-input widget values back into the companies list."""
    for i in range(len(st.session_state.companies)):
        key = f"co_{i}"
        if key in st.session_state:
            st.session_state.companies[i] = st.session_state[key]


def _add_company() -> None:
    _sync_inputs()
    st.session_state.companies.append("")


def _remove_company(idx: int) -> None:
    _sync_inputs()
    st.session_state.companies.pop(idx)


def _clear_all() -> None:
    st.session_state.companies = [""]


# ─────────────────────────────────────────────────────────────────────────────
# Header
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <h1 style='text-align:center;color:#f5f3f4;font-size:40px;margin-bottom:4px'>
    📊 Bottom-Up Beta Calculator
    </h1>

    <p style='text-align:center;color:#64748b;font-size:14px;margin-bottom:8px'>
    Historical closing price extraction · Yahoo Finance · Consolidated Excel export
    </p>

    <p style='text-align:center;font-size:14px'>
    Created by
    <a href='https://www.linkedin.com/in/srijanchiripal/'
       target='_blank'
       style='text-decoration:none;font-weight:600;color:#41ead4'>
       Srijan Chiripal
    </a>
    </p>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    """<div class="note">
⚠️ <strong>Ticker Symbols — Important:</strong>
Enter tickers <em>exactly</em> as shown on Yahoo Finance.<br>
Examples: &nbsp;
<code>AAPL</code> (Apple USA) &nbsp;·&nbsp;
<code>MSFT</code> (Microsoft) &nbsp;·&nbsp;
<code>RELIANCE.NS</code> (Reliance NSE India) &nbsp;·&nbsp;
<code>TCS.NS</code> (TCS NSE) &nbsp;·&nbsp;
<code>0700.HK</code> (Tencent HK) &nbsp;·&nbsp;
<code>SAP.DE</code> (SAP Frankfurt) &nbsp;·&nbsp;
<code>BP.L</code> (BP London).<br>
Not sure? Visit <a href="https://finance.yahoo.com" target="_blank">finance.yahoo.com</a>,
search the company name, and copy the ticker symbol shown in the URL or page header.
The <strong>benchmark index</strong> is selected from the dropdown — you do not type it.
</div>""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# Section 1 — Companies
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="card"><div class="card-title">🏢 Section 1 — Companies</div>',
    unsafe_allow_html=True,
)

btn_add, btn_clr, _ = st.columns([1.2, 1.2, 7.6])
with btn_add:
    st.button("➕ Add company", on_click=_add_company)
with btn_clr:
    st.button("🗑️ Clear all", on_click=_clear_all)

n_co = len(st.session_state.companies)
for row_start in range(0, n_co, 3):
    cols = st.columns(3)
    for col_offset in range(3):
        co_idx = row_start + col_offset
        if co_idx >= n_co:
            break
        with cols[col_offset]:
            inner_l, inner_r = st.columns([9, 1])
            with inner_l:
                st.text_input(
                    f"Company {co_idx + 1}",
                    value=st.session_state.companies[co_idx],
                    placeholder="e.g. AAPL / RELIANCE.NS",
                    key=f"co_{co_idx}",
                )
            with inner_r:
                if n_co > 1:
                    st.markdown(
                        "<div style='margin-top:28px'></div>",
                        unsafe_allow_html=True,
                    )
                    st.button(
                        "✖",
                        key=f"rm_{co_idx}",
                        on_click=_remove_company,
                        args=(co_idx,),
                        help="Remove this company",
                    )

st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Section 2 — Valuation Parameters
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="card"><div class="card-title">⚙️ Section 2 — Valuation Parameters</div>',
    unsafe_allow_html=True,
)

col1, col2, col3, col4 = st.columns([2.8, 1.5, 1.5, 1.5])

with col1:
    idx_name: str = st.selectbox("📈 Benchmark Index", list(INDICES.keys()))
    if INDICES[idx_name] == "__CUSTOM__":
        custom_sym = st.text_input(
            "Custom index ticker", placeholder="e.g. ^STOXX  or  000001.SS"
        )
        idx_ticker: str = custom_sym.strip().upper()
        idx_label: str  = idx_ticker or "Custom Index"
    else:
        idx_ticker = INDICES[idx_name]
        # strip trailing parenthetical for a clean label
        idx_label = idx_name.split("(")[0].strip(" ─")

with col2:
    val_date: date = st.date_input(
        "📅 Valuation Date",
        value=date.today(),
        max_value=date.today(),
    )

with col3:
    yrs: int = st.number_input(
        "📆 Number of Years", min_value=1, max_value=30, value=5, step=1
    )

with col4:
    freq: str = st.radio("🔄 Frequency", ["Daily", "Weekly", "Monthly"], index=2)

st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SEARCH button
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("---")
go = st.button(
    "🔍  FETCH DATA & GENERATE EXCEL",
    use_container_width=True,
    type="primary",
)

# ─────────────────────────────────────────────────────────────────────────────
# Processing
# ─────────────────────────────────────────────────────────────────────────────
if go:
    _sync_inputs()
    tickers: list[str] = [
        t.strip().upper() for t in st.session_state.companies if t.strip()
    ]

    if not tickers:
        st.error("⚠️  Please enter at least one company ticker.")
        st.stop()
    if not idx_ticker or idx_ticker.startswith("──"):
        st.error("⚠️  Please specify a benchmark index ticker.")
        st.stop()

    end_dt   = pd.Timestamp(val_date)
    start_dt = end_dt - relativedelta(years=int(yrs))
    interval = INTERVAL[freq]
    expected = int(yrs * EST_PTS[freq])

    data_series: dict[str, pd.Series] = {}
    error_log:   list[dict]            = []
    counts:      dict[str, int]        = {}

    prog = st.progress(0.0, text="Initialising …")
    stat = st.empty()
    total_items = len(tickers) + 1   # companies + index

    # ── Helper: download closing prices for one symbol ────────────────────────
    def fetch_one(sym: str) -> "pd.Series | None":
        common_kwargs = dict(
            start    = start_dt,
            end      = end_dt + pd.Timedelta(days=1),
            interval = interval,
            progress = False,
            auto_adjust = True,
            actions  = False,
        )
        # yfinance ≥ 0.2.55 supports multi_level_index=False for flat columns
        try:
            raw = yf.download(sym, **common_kwargs, multi_level_index=False)
        except TypeError:
            raw = yf.download(sym, **common_kwargs)

        if raw is None or raw.empty:
            return None

        # Flatten MultiIndex columns if present (older yfinance versions)
        if isinstance(raw.columns, pd.MultiIndex):
            lvl0 = raw.columns.get_level_values(0)
            if "Close" in lvl0:
                close = raw["Close"]
                if isinstance(close, pd.DataFrame):
                    close = close.iloc[:, 0]
            else:
                return None
        else:
            if "Close" not in raw.columns:
                return None
            close = raw["Close"]

        close = close.dropna()
        return close if not close.empty else None

    # ── Fetch index ───────────────────────────────────────────────────────────
    stat.info(f"Fetching benchmark index: {idx_label} ({idx_ticker}) …")
    try:
        s = fetch_one(idx_ticker)
        if s is None:
            error_log.append({
                "Symbol": idx_ticker,
                "Role":   "Index",
                "Issue":  "No data returned. Please verify the ticker is correct.",
            })
        else:
            s.name = f"{idx_label} ({idx_ticker})"
            data_series["__INDEX__"] = s
            counts["__INDEX__"]      = len(s)
    except Exception as exc:
        error_log.append({"Symbol": idx_ticker, "Role": "Index", "Issue": str(exc)})

    prog.progress(1 / total_items, text=f"Index done. Fetching companies …")

    # ── Fetch companies ───────────────────────────────────────────────────────
    for i, tk in enumerate(tickers):
        stat.info(f"Fetching {tk}  ({i + 1} / {len(tickers)}) …")
        try:
            s = fetch_one(tk)
            if s is None:
                error_log.append({
                    "Symbol": tk,
                    "Role":   "Company",
                    "Issue":  "No data found. Ticker may be incorrect, delisted, "
                              "or data unavailable for the selected period.",
                })
            else:
                s.name = tk
                data_series[tk] = s
                counts[tk]      = len(s)
        except Exception as exc:
            error_log.append({"Symbol": tk, "Role": "Company", "Issue": str(exc)})

        prog.progress((i + 2) / total_items, text=f"Fetched {i + 1}/{len(tickers)} companies …")

    prog.progress(1.0, text="Building Excel workbook …")

    if not data_series:
        stat.error(
            "❌  No data could be retrieved. "
            "Please verify your tickers and try again."
        )
        st.stop()

    # ── Combine into one aligned DataFrame ────────────────────────────────────
    frames: list[pd.Series] = []
    if "__INDEX__" in data_series:
        frames.append(data_series.pop("__INDEX__"))    # index first
    for s in data_series.values():
        frames.append(s)

    df: pd.DataFrame = pd.concat(frames, axis=1).sort_index()
    df.index          = pd.to_datetime(df.index)
    df.index.name     = "Date"

    idx_count: int = counts.get("__INDEX__", 0)

    # ═══════════════════════════════════════════════════════════════
    # Excel builder
    # ═══════════════════════════════════════════════════════════════

    # ── Style helpers ──────────────────────────────────────────────
    DARK  = "0F2744"
    BLUE  = "1A5276"
    MID   = "2E86C1"
    LTBLU = "EAF4FB"
    WHITE = "FFFFFF"
    ALT   = "F4F8FC"
    GREEN = "D5F5E3"
    YELL  = "FEF9E7"
    RED   = "FADBD8"
    DKRED = "C0392B"
    DKGRN = "1D6F42"

    def xfill(hex_: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_)

    def xfont(bold=False, color="111111", size=10) -> Font:
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def xborder() -> Border:
        side = Side(style="thin", color="D0D7DE")
        return Border(left=side, right=side, top=side, bottom=side)

    al_c = Alignment(horizontal="center", vertical="center")
    al_l = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    al_r = Alignment(horizontal="right",  vertical="center")

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1 — Historical Closing Prices
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Historical Closing Prices"

    n_cols  = len(df.columns)
    last_cl = get_column_letter(n_cols + 1)

    # Row 1 — banner
    ws1.merge_cells(f"A1:{last_cl}1")
    c1 = ws1["A1"]
    c1.value = (
        f"Historical Closing Prices   ·   Frequency: {freq}   ·   "
        f"Period: {yrs} Year(s)   ·   Valuation Date: {val_date}"
    )
    c1.font      = xfont(bold=True, color=WHITE, size=12)
    c1.fill      = xfill(DARK)
    c1.alignment = al_c
    ws1.row_dimensions[1].height = 26

    # Row 2 — column headers
    def make_hdr(ws, row, col, val, hex_color):
        c = ws.cell(row, col, val)
        c.font = xfont(bold=True, color=WHITE)
        c.fill = xfill(hex_color)
        c.alignment = al_c
        c.border = xborder()
        return c

    make_hdr(ws1, 2, 1, "Date", DARK)
    ws1.column_dimensions["A"].width = 14
    ws1.row_dimensions[2].height = 20

    for ci, col_name in enumerate(df.columns, start=2):
        # First column after Date is always the index → darker blue
        color = BLUE if ci == 2 else MID
        make_hdr(ws1, 2, ci, col_name, color)
        ws1.column_dimensions[get_column_letter(ci)].width = max(
            len(col_name) + 4, 18
        )

    # Data rows
    for ri, (dt_idx, row_vals) in enumerate(df.iterrows()):
        er    = ri + 3
        rfill = xfill(LTBLU) if ri % 2 == 0 else xfill(WHITE)

        dc = ws1.cell(er, 1, dt_idx.date())
        dc.number_format = "YYYY-MM-DD"
        dc.font          = xfont(bold=True)
        dc.fill          = rfill
        dc.alignment     = al_c
        dc.border        = xborder()

        for ci, v in enumerate(row_vals, start=2):
            val = None if (hasattr(v, "__class__") and str(type(v)) == "<class 'float'>"
                          and str(v) == "nan") or (
                          hasattr(pd, "isna") and pd.isna(v)) else round(float(v), 4)
            vc = ws1.cell(er, ci, val)
            vc.number_format = "#,##0.0000"
            vc.font          = xfont()
            vc.fill          = rfill
            vc.alignment     = al_r
            vc.border        = xborder()

    ws1.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2 — Error & QC Report
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Error & QC Report")

    for cl, w in zip("ABCDEFGH", [5, 32, 12, 15, 18, 15, 20, 38]):
        ws2.column_dimensions[cl].width = w

    # Banner
    ws2.merge_cells("A1:H1")
    c = ws2["A1"]
    c.value     = "Error & Quality Control Report"
    c.font      = xfont(bold=True, color=WHITE, size=13)
    c.fill      = xfill(DARK)
    c.alignment = al_c
    ws2.row_dimensions[1].height = 28

    def set_cell(ws, row, col, val, fnt=None, fl=None, al=None, brd=True):
        c = ws.cell(row, col, val)
        if fnt: c.font      = fnt
        if fl:  c.fill      = fl
        if al:  c.alignment = al
        if brd: c.border    = xborder()
        return c

    r = 3  # current row pointer

    # ── A: Run parameters ─────────────────────────────────────────────────────
    ws2.cell(r, 1, "A — RUN PARAMETERS").font = xfont(bold=True, color=BLUE, size=11)
    r += 1

    params_tbl = [
        ("Valuation Date",       str(val_date)),
        ("Period (Years)",       str(yrs)),
        ("Data Frequency",       freq),
        ("Expected Data Points", f"~ {expected}  (approx. trading periods)"),
        ("Benchmark Index",      f"{idx_label}  ({idx_ticker})"),
        ("Companies Requested",  str(len(tickers))),
        ("Companies Retrieved",  str(len([t for t in tickers if t in counts]))),
        ("Report Generated",     datetime.now().strftime("%Y-%m-%d  %H:%M:%S")),
    ]
    for lbl, val in params_tbl:
        set_cell(ws2, r, 1, lbl,
                 fnt=xfont(bold=True), fl=xfill(LTBLU), al=al_l)
        set_cell(ws2, r, 2, val,
                 fnt=xfont(),          fl=xfill(WHITE),  al=al_l)
        ws2.merge_cells(f"B{r}:H{r}")
        r += 1

    r += 1  # blank spacer

    # ── B: Data count / deviation table ───────────────────────────────────────
    ws2.cell(r, 1, "B — DATA COUNT & DEVIATION FLAGS").font = xfont(
        bold=True, color=BLUE, size=11
    )
    r += 1

    cnt_hdrs = [
        "#", "Symbol / Name", "Role",
        "Data Points", "Reference", "Deviation (%)", "Status", "Note",
    ]
    for ci, hdr in enumerate(cnt_hdrs, start=1):
        set_cell(ws2, r, ci, hdr,
                 fnt=xfont(bold=True, color=WHITE),
                 fl=xfill(BLUE), al=al_c)
    ws2.row_dimensions[r].height = 18
    r += 1

    def write_count_row(ws, rw, num, symbol, role, cnt, ref):
        dev = ((cnt - ref) / ref * 100) if ref > 0 else None
        if cnt == 0:
            status = "❌ NO DATA"
            note   = "No data was retrieved for this symbol."
            fhex   = RED
        elif dev is not None and abs(dev) > 10:
            status = "⚠️ DEVIATION"
            note   = f"Count deviates by {dev:.1f}% from the reference value."
            fhex   = YELL
        else:
            status = "✅ OK"
            note   = "Data count is within acceptable range (±10% of reference)."
            fhex   = GREEN

        row_vals = [
            num, symbol, role, cnt, ref,
            f"{dev:.1f}%" if dev is not None else "N/A",
            status, note,
        ]
        for ci, v in enumerate(row_vals, start=1):
            set_cell(ws, rw, ci, v,
                     fnt=xfont(bold=(ci == 2)),
                     fl=xfill(fhex),
                     al=al_l if ci in (2, 8) else al_c)

    # Index row  (reference = expected based on freq × years)
    write_count_row(ws2, r, 1,
                    f"{idx_label} ({idx_ticker})", "Index",
                    idx_count, expected)
    r += 1

    # Company rows  (reference = index data count for alignment)
    for j, tk in enumerate(tickers, start=2):
        cnt = counts.get(tk, 0)
        write_count_row(ws2, r, j, tk, "Company", cnt, idx_count)
        r += 1

    r += 1  # spacer

    # ── C: Error log ──────────────────────────────────────────────────────────
    ws2.cell(r, 1, "C — ERROR LOG").font = xfont(bold=True, color=BLUE, size=11)
    r += 1

    if error_log:
        err_hdrs = ["#", "Symbol", "Role", "Issue / Error Description"]
        for ci, hdr in enumerate(err_hdrs, start=1):
            set_cell(ws2, r, ci, hdr,
                     fnt=xfont(bold=True, color=WHITE),
                     fl=xfill(DKRED), al=al_c)
        r += 1
        for j, err in enumerate(error_log, start=1):
            for ci, v in enumerate(
                [j, err["Symbol"], err["Role"], err["Issue"]], start=1
            ):
                set_cell(ws2, r, ci, v,
                         fnt=xfont(),
                         fl=xfill(RED),
                         al=al_l if ci == 4 else al_c)
            r += 1
    else:
        c = set_cell(ws2, r, 1,
                     "✅  No errors encountered — all tickers fetched successfully.",
                     fnt=xfont(bold=True, color=DKGRN),
                     fl=xfill(GREEN), al=al_l)
        ws2.merge_cells(f"A{r}:H{r}")
        r += 1

    ws2.freeze_panes = "A2"

    # ── Save to buffer ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    prog.empty()
    stat.empty()

    # ─────────────────────────────────────────────────────────────────────────
    # Results UI
    # ─────────────────────────────────────────────────────────────────────────
    n_ok = len([t for t in tickers if t in counts])
    st.success(
        f"✅  **{len(df):,} rows** retrieved  ·  "
        f"**{n_ok} / {len(tickers)} companies** successful  ·  "
        f"Index: **{idx_label}**  ·  Frequency: **{freq}**"
    )

    if error_log:
        st.warning(
            f"⚠️  **{len(error_log)} issue(s)** encountered — "
            "see the *Error & QC Report* sheet in the Excel file for details."
        )

    with st.expander("📋  Data Preview — first 20 rows"):
        preview = df.head(20).copy()
        preview.index = preview.index.date
        st.dataframe(
            preview.style.format("{:.4f}", na_rep="—"),
            use_container_width=True,
        )

    with st.expander("🔢  Data Count Summary"):
        count_rows = []
        for sym, cnt in {**{"__INDEX__": idx_count}, **counts}.items():
            if sym == "__INDEX__":
                label_ = f"{idx_label} ({idx_ticker})"
                ref_   = expected
            else:
                label_ = sym
                ref_   = idx_count
            dev_ = f"{(cnt - ref_) / ref_ * 100:.1f}%" if ref_ > 0 else "N/A"
            flag = "✅ OK" if cnt > 0 and (ref_ == 0 or abs((cnt - ref_) / ref_ * 100) <= 10) \
                   else ("⚠️ Deviation" if cnt > 0 else "❌ No Data")
            count_rows.append({
                "Symbol":       label_,
                "Data Points":  cnt,
                "Reference":    ref_,
                "Deviation (%)": dev_,
                "Status":       flag,
            })
        st.dataframe(pd.DataFrame(count_rows), use_container_width=True)

    if error_log:
        with st.expander("❌  Error Details"):
            st.dataframe(pd.DataFrame(error_log), use_container_width=True)

    # ── Download button ────────────────────────────────────────────────────────
    fname = f"beta_closing_prices_{val_date}_{freq.lower()}.xlsx"
    st.download_button(
        label="⬇️  Download Excel Report",
        data=buf,
        file_name=fname,
        mime=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        use_container_width=True,
        type="primary",
    )
